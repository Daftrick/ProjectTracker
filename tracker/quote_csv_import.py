import csv
import datetime
import io

from .catalog import catalog_name_key

_ENCODING_ERROR = (
    "El CSV tiene codificación no compatible (posiblemente ANSI/cp1252). "
    "Verifica que AutoCAD pudo escribir el archivo en UTF-8. "
    "Si el problema persiste, abre el CSV en un editor de texto y guárdalo como UTF-8."
)

_XLS_LEGACY_ERROR = (
    "El archivo es un Excel antiguo (.xls). Ábrelo en Excel y guárdalo como "
    ".xlsx o como CSV (UTF-8) para poder importarlo."
)

_OPENPYXL_MISSING_ERROR = (
    "El archivo es un Excel (.xlsx) pero openpyxl no está instalado. "
    "Reinicia la app con INICIAR.bat para instalar dependencias."
)


QUOTE_DESCRIPTION_COLUMNS = (
    "description",
    "descripcion",
    "descripción",
    "nombre",
    "concepto",
    "partida",
    "item",
    "articulo",
    "artículo",
)
QUOTE_UNIT_COLUMNS = ("unit", "unidad", "u")
QUOTE_QTY_COLUMNS = ("qty", "cantidad", "cant", "quantity")
QUOTE_PRICE_COLUMNS = (
    "price",
    "precio",
    "precio_unitario",
    "precio unitario",
    "unit_price",
    "pu",
    "p.u.",
)
QUOTE_SECTION_COLUMNS = ("section", "seccion", "sección", "categoria", "categoría", "grupo")


def _clean(value):
    return str(value or "").strip()


def _header_key(value):
    return _clean(value).lower().lstrip("\ufeff")


def _read_sample(path):
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            return handle.read(4096)
    except UnicodeDecodeError:
        return None


def _detect_dialect(path):
    sample = _read_sample(path)
    if not sample:
        return csv.excel
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        return csv.excel


def _parse_float(value, row_number, field_label, errors, required=False, default=0.0):
    raw = _clean(value).replace(",", ".")
    if not raw:
        if required:
            errors.append(f"Fila {row_number}: {field_label} es requerida.")
        return default
    try:
        return float(raw)
    except ValueError:
        errors.append(f"Fila {row_number}: {field_label} debe ser un numero valido.")
        return default


def _metadata_value(row):
    values = [_clean(value) for value in row]
    return next((value for value in values[1:] if value), "")


def _build_catalog_index(catalog):
    index = {}
    for item in catalog or []:
        key = catalog_name_key(item.get("nombre", ""))
        if key and key not in index:
            index[key] = item
    return index


def _match_catalog(description, index):
    return index.get(catalog_name_key(description))


def _column_index(headers, aliases):
    for index, header in enumerate(headers):
        if _header_key(header) in aliases:
            return index
    return None


def _row_value(row, index):
    if index is None or index >= len(row):
        return ""
    return _clean(row[index])


def _find_header_row(rows, metadata):
    for index, row in enumerate(rows):
        if not any(_clean(value) for value in row):
            continue
        first = _clean(row[0])
        if first.startswith("#"):
            metadata[first.lstrip("#").strip().lower()] = _metadata_value(row)
            continue
        return index
    return None


def parse_quote_csv(path, catalog=None):
    """Parse a LISP-exported client quote CSV into quote draft data."""
    dialect = _detect_dialect(path)
    catalog_index = _build_catalog_index(catalog)
    metadata = {}
    items = []
    errors = []
    warnings = []

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, dialect=dialect))
    except UnicodeDecodeError:
        return {"items": [], "metadata": {}, "errors": [_ENCODING_ERROR], "warnings": []}

    header_row = _find_header_row(rows, metadata)
    if header_row is None:
        return {"items": [], "metadata": metadata, "errors": ["El CSV no tiene encabezados."], "warnings": []}

    headers = [_header_key(value) for value in rows[header_row]]
    description_index = _column_index(headers, QUOTE_DESCRIPTION_COLUMNS)
    unit_index = _column_index(headers, QUOTE_UNIT_COLUMNS)
    qty_index = _column_index(headers, QUOTE_QTY_COLUMNS)
    price_index = _column_index(headers, QUOTE_PRICE_COLUMNS)
    section_index = _column_index(headers, QUOTE_SECTION_COLUMNS)

    if description_index is None:
        errors.append("El CSV debe incluir una columna description o descripcion.")
    if qty_index is None:
        errors.append("El CSV debe incluir una columna qty o cantidad.")
    if errors:
        return {"items": [], "metadata": metadata, "errors": errors, "warnings": warnings}

    seen = {}
    for zero_based_index, row in enumerate(rows[header_row + 1:], start=header_row + 1):
        row_number = zero_based_index + 1
        if not any(_clean(value) for value in row):
            continue
        first = _clean(row[0])
        if first.startswith("#"):
            metadata[first.lstrip("#").strip().lower()] = _metadata_value(row)
            continue

        description = _row_value(row, description_index)
        unit = _row_value(row, unit_index) or "pza"
        qty = _parse_float(_row_value(row, qty_index), row_number, "cantidad", errors, required=True)
        price = _parse_float(_row_value(row, price_index), row_number, "precio unitario", errors, default=0.0)
        section = _row_value(row, section_index)

        if not description:
            errors.append(f"Fila {row_number}: descripcion es requerida.")
            continue
        if qty <= 0:
            errors.append(f"Fila {row_number}: cantidad debe ser mayor a 0.")
        if price < 0:
            errors.append(f"Fila {row_number}: precio unitario no puede ser negativo.")

        catalog_item = _match_catalog(description, catalog_index)
        key = (catalog_name_key(description), unit.lower())
        seen[key] = seen.get(key, 0) + 1
        items.append({
            "description": description,
            "unit": unit,
            "qty": qty,
            "price": price,
            "total": round(qty * price, 2),
            "csv_row_number": row_number,
            "catalog_item_id": catalog_item.get("id", "") if catalog_item else "",
            "catalog_description": catalog_item.get("descripcion", "") if catalog_item else "",
            "section": section,
            "origen": "csv",
        })

    duplicate_count = sum(1 for count in seen.values() if count > 1)
    if duplicate_count:
        warnings.append(f"{duplicate_count} concepto(s) aparecen mas de una vez con la misma unidad.")
    if not items and not errors:
        errors.append("El CSV no contiene partidas para importar.")
    return {"items": items, "metadata": metadata, "errors": errors, "warnings": warnings}


# ── Soporte para archivos Excel (.xlsx) ────────────────────────────────────────
# Algunos usuarios suben la cotización exportada como Excel (a veces con la
# extensión cambiada a .csv). Detectamos el contenido real y lo parseamos.

_XLSX_METADATA_LABELS = {
    "fecha": "fecha",
    "moneda": "currency",
    "cliente": "cliente",
    "proyecto": "proyecto",
    "cotizacion": "cotizacion",
    "cotización": "cotizacion",
    "version": "version",
    "versión": "version",
}


def _cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _xlsx_rows(path):
    """Lee la primera hoja de un .xlsx como lista de filas (listas de texto)."""
    from openpyxl import load_workbook  # import perezoso

    with open(path, "rb") as handle:
        data = handle.read()
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [[_cell_to_text(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _find_table_header(rows):
    """Localiza la fila de encabezados buscando columnas reconocibles."""
    for index, row in enumerate(rows):
        headers = [_header_key(value) for value in row]
        has_description = any(header in QUOTE_DESCRIPTION_COLUMNS for header in headers)
        has_qty = any(header in QUOTE_QTY_COLUMNS for header in headers)
        if has_description and has_qty:
            return index, headers
    return None, []


def _xlsx_metadata(rows, header_index):
    metadata = {}
    for row in rows[:header_index]:
        if not any(_clean(value) for value in row):
            continue
        label = _header_key(row[0] if row else "").rstrip(":").strip()
        value = _metadata_value(row)
        if label and value:
            metadata[_XLSX_METADATA_LABELS.get(label, label)] = value
    return metadata


def parse_quote_xlsx(path, catalog=None):
    """Parsea una cotización exportada como Excel hacia datos de borrador.

    Reconoce el formato de exportación de la app (filas informativas, encabezado
    de tabla, subtítulos de sección y filas de subtotal/IVA/total) y extrae solo
    las partidas reales (las que tienen cantidad).
    """
    try:
        rows = _xlsx_rows(path)
    except ImportError:
        return {"items": [], "metadata": {}, "errors": [_OPENPYXL_MISSING_ERROR], "warnings": []}
    except Exception as exc:  # archivo corrupto o no legible
        return {
            "items": [],
            "metadata": {},
            "errors": [f"No se pudo leer el Excel: {exc}"],
            "warnings": [],
        }

    header_index, headers = _find_table_header(rows)
    if header_index is None:
        return {
            "items": [],
            "metadata": {},
            "errors": ["El Excel no tiene una tabla con columnas reconocibles (Nombre/Descripción y Cantidad)."],
            "warnings": [],
        }

    metadata = _xlsx_metadata(rows, header_index)
    catalog_index = _build_catalog_index(catalog)
    description_index = _column_index(headers, QUOTE_DESCRIPTION_COLUMNS)
    unit_index = _column_index(headers, QUOTE_UNIT_COLUMNS)
    qty_index = _column_index(headers, QUOTE_QTY_COLUMNS)
    price_index = _column_index(headers, QUOTE_PRICE_COLUMNS)
    section_index = _column_index(headers, QUOTE_SECTION_COLUMNS)

    items = []
    errors = []
    warnings = []
    seen = {}

    for zero_based_index, row in enumerate(rows[header_index + 1:], start=header_index + 1):
        row_number = zero_based_index + 1
        if not any(_clean(value) for value in row):
            continue

        description = _row_value(row, description_index)
        unit = _row_value(row, unit_index)
        qty_raw = _row_value(row, qty_index)
        price_raw = _row_value(row, price_index)
        section = _row_value(row, section_index)

        # Filas sin cantidad: subtítulos de sección, subtotales o totales. Se
        # omiten, salvo que parezcan una partida real (con unidad o precio) a la
        # que le falta la cantidad: en ese caso es un error que conviene avisar.
        if not qty_raw:
            if description and (unit or price_raw):
                errors.append(f"Fila {row_number}: cantidad es requerida.")
            continue

        if not description:
            errors.append(f"Fila {row_number}: descripcion es requerida.")
            continue

        qty = _parse_float(qty_raw, row_number, "cantidad", errors, required=True)
        price = _parse_float(price_raw, row_number, "precio unitario", errors, default=0.0)
        if qty <= 0:
            errors.append(f"Fila {row_number}: cantidad debe ser mayor a 0.")
        if price < 0:
            errors.append(f"Fila {row_number}: precio unitario no puede ser negativo.")

        unit = unit or "pza"
        catalog_item = _match_catalog(description, catalog_index)
        key = (catalog_name_key(description), unit.lower())
        seen[key] = seen.get(key, 0) + 1
        items.append({
            "description": description,
            "unit": unit,
            "qty": qty,
            "price": price,
            "total": round(qty * price, 2),
            "csv_row_number": row_number,
            "catalog_item_id": catalog_item.get("id", "") if catalog_item else "",
            "catalog_description": catalog_item.get("descripcion", "") if catalog_item else "",
            "section": section,
            "origen": "csv",
        })

    duplicate_count = sum(1 for count in seen.values() if count > 1)
    if duplicate_count:
        warnings.append(f"{duplicate_count} concepto(s) aparecen mas de una vez con la misma unidad.")
    if not items and not errors:
        errors.append("El Excel no contiene partidas para importar.")
    return {"items": items, "metadata": metadata, "errors": errors, "warnings": warnings}


def _sniff_filetype(path):
    """Detecta el tipo real del archivo por sus bytes iniciales (no por extensión)."""
    try:
        with open(path, "rb") as handle:
            head = handle.read(8)
    except OSError:
        return "csv"
    if head[:4] == b"PK\x03\x04":  # ZIP -> .xlsx
        return "xlsx"
    if head[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 -> .xls antiguo
        return "xls"
    return "csv"


def parse_quote_file(path, catalog=None):
    """Parsea una cotización desde CSV o Excel, según el contenido real del archivo.

    El usuario puede subir un .csv real, un .xlsx, o incluso un Excel con la
    extensión cambiada a .csv. Se decide por los bytes mágicos, no por el nombre.
    """
    kind = _sniff_filetype(path)
    if kind == "xlsx":
        return parse_quote_xlsx(path, catalog=catalog)
    if kind == "xls":
        return {"items": [], "metadata": {}, "errors": [_XLS_LEGACY_ERROR], "warnings": []}
    return parse_quote_csv(path, catalog=catalog)
