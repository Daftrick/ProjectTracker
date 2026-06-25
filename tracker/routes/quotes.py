import copy
import os

from io import BytesIO

from flask import Blueprint, Response, flash, redirect, render_template, request, send_file, url_for

from ..catalog import aggregate_quote_items, approve_quote, catalog_maps, hydrate_quote, next_quote_number, quote_type_key, safe_float
from ..csv_catalog_validation import validate_csv_catalog_items
from ..deletions import purge_deleted_catalog_items_from_record
from ..form_models import quote_default_numbers, quote_from_form
from ..pdfs import build_quote_pdf
from ..quote_csv_import import parse_quote_file
from ..auth import admin_required
from ..storage import load, new_id, save, today
from ..validators import validate_quote_form

bp = Blueprint("quotes_bp", __name__)


def _render_quote_form(project, quote, quotes, field_errors=None, quote_id=None, form_action=None):
    from ..quote_templates_config import get_quote_templates
    return render_template(
        "quote_project_form.html",
        project=project,
        quote=quote,
        today=today(),
        field_errors=field_errors or {},
        form_action=form_action,
        quote_templates=get_quote_templates(),
        **quote_default_numbers(project, quotes, quote_id=quote_id),
    )


def _find_project(project_id):
    return next((item for item in load("projects") if item["id"] == project_id), None)


def _flash_csv_catalog_errors(validation, label="CSV COT"):
    errors = validation.get("errors", [])
    flash(
        f"{label} no se puede importar: {len(errors)} partida(s) no pasan validación contra catálogo.",
        "danger",
    )
    for error in errors[:10]:
        flash(error, "warning")
    if len(errors) > 10:
        flash(f"Se omitieron {len(errors) - 10} error(es) adicional(es).", "warning")


def _quote_preview_from_csv(project, parsed, filename, quotes):
    metadata = parsed.get("metadata", {})
    quote_type = quote_type_key(metadata.get("quote_type") or metadata.get("tipo") or "Proyecto")
    date_str = metadata.get("fecha") or metadata.get("date") or today()
    subtotal = round(sum(safe_float(item.get("total", 0)) for item in parsed.get("items", [])), 2)
    tax_rate = 16.0
    notes_parts = [f"Importada desde CSV: {filename}"]
    if metadata.get("source"):
        notes_parts.append(f"Origen: {metadata['source']}")
    if metadata.get("drawing"):
        notes_parts.append(f"Dibujo: {metadata['drawing']}")
    return hydrate_quote(
        {
            "is_import_preview": True,
            "project_id": project["id"],
            "quote_number": next_quote_number(project, quotes, quote_type, date_str),
            "quote_type": quote_type,
            "version": metadata.get("version") or project.get("version", ""),
            "client": project.get("client", ""),
            "project_name": project.get("name", ""),
            "date": date_str,
            "valid_until": "",
            "items": parsed.get("items", []),
            "subtotal": subtotal,
            "tax_rate": tax_rate,
            "tax": round(subtotal * tax_rate / 100, 2),
            "total": round(subtotal + subtotal * tax_rate / 100, 2),
            "currency": metadata.get("currency") or "MXN",
            "notes": "\n".join(notes_parts),
            "project_basis_note": "",
            "csv_origen": filename,
            "csv_metadata": metadata,
        },
        *catalog_maps(),
    )


@bp.route("/projects/<project_id>/quote/new", methods=["GET", "POST"], endpoint="new_quote")
def new_quote(project_id):
    project = _find_project(project_id)
    if not project:
        return redirect(url_for("dashboard"))
    quotes = load("quotes")
    if request.method == "POST":
        validation = validate_quote_form(request.form)
        if not validation["ok"]:
            for error in validation["errors"]:
                flash(error, "warning")
            return _render_quote_form(
                project,
                quote_from_form(request.form),
                quotes,
                field_errors=validation["field_errors"],
            )
        quote_type = validation["quote_type"]
        date_str = validation["date"]
        quote_number = request.form.get("quote_number", "").strip() or next_quote_number(project, quotes, quote_type, date_str)
        quote = {
            "id": new_id(),
            "project_id": project_id,
            "quote_number": quote_number,
            "quote_type": quote_type,
            "version": validation["version"],
            "client": project.get("client", ""),
            "project_name": project.get("name", ""),
            "date": date_str,
            "valid_until": validation["valid_until"],
            "items": validation["items"],
            "subtotal": validation["subtotal"],
            "tax_rate": validation["tax_rate"],
            "tax": round(validation["subtotal"] * validation["tax_rate"] / 100, 2),
            "total": round(validation["subtotal"] + validation["subtotal"] * validation["tax_rate"] / 100, 2),
            "currency": validation["currency"],
            "notes": validation["notes"],
            "project_basis_note": validation["project_basis_note"] if quote_type == "Extraordinaria" else "",
            "specs": validation["specs"],
            "created_at": today(),
        }
        csv_origen = request.form.get("csv_origen", "").strip()
        if csv_origen:
            quote["csv_origen"] = csv_origen
            quote["csv_sources"] = [csv_origen]
            metadata_keys = request.form.getlist("csv_metadata_key[]")
            metadata_values = request.form.getlist("csv_metadata_value[]")
            quote["csv_metadata"] = {
                key.strip(): metadata_values[index].strip()
                for index, key in enumerate(metadata_keys)
                if key.strip() and index < len(metadata_values) and metadata_values[index].strip()
            }
        quotes.append(quote)
        save("quotes", quotes)
        flash(f"Cotización {quote['quote_number']} creada.", "success")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote["id"]))
    return _render_quote_form(project, None, quotes)


@bp.route("/projects/<project_id>/quote/import", methods=["POST"], endpoint="import_quote_csv")
def import_quote_csv(project_id):
    project = _find_project(project_id)
    if not project:
        return redirect(url_for("dashboard"))
    if project.get("closed_at"):
        flash("El proyecto esta cerrado. Reabrelo para importar una cotizacion.", "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    uploaded = request.files.get("quote_csv")
    if not uploaded or not uploaded.filename:
        flash("Selecciona un CSV COT para importar.", "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    clean_name = os.path.basename(uploaded.filename)
    if not clean_name.lower().endswith((".csv", ".xlsx", ".xls")):
        flash("El archivo debe ser CSV o Excel (.csv, .xlsx).", "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    import tempfile

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as handle:
            temp_path = handle.name
            uploaded.save(handle)
        catalog = load("catalogo")
        parsed = parse_quote_file(temp_path, catalog=catalog)
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)

    if parsed["errors"]:
        for error in parsed["errors"]:
            flash(error, "warning")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    catalog_validation = validate_csv_catalog_items(parsed["items"], catalog, kind="COT")
    if not catalog_validation["ok"]:
        _flash_csv_catalog_errors(catalog_validation, "CSV COT")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    metadata = parsed.get("metadata", {})
    if metadata.get("proyecto_clave") and metadata["proyecto_clave"] != project.get("clave"):
        flash("La clave de proyecto indicada en el CSV no coincide con este proyecto.", "warning")
    missing_catalog = [item for item in parsed["items"] if not item.get("catalog_item_id")]
    if missing_catalog:
        flash(f"{len(missing_catalog)} partida(s) no tienen coincidencia exacta en catalogo.", "warning")
    for warning in parsed.get("warnings", []):
        flash(warning, "warning")

    quotes = load("quotes")
    preview = _quote_preview_from_csv(project, parsed, clean_name, quotes)
    return _render_quote_form(
        project,
        preview,
        quotes,
        form_action=url_for("quotes_bp.new_quote", project_id=project_id),
    )


@bp.route("/projects/<project_id>/quote/<quote_id>/edit", methods=["GET", "POST"], endpoint="edit_quote")
def edit_quote(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id), None)
    if not project or not quote:
        return redirect(url_for("project_detail", project_id=project_id))
    if request.method == "POST":
        validation = validate_quote_form(request.form)
        if not validation["ok"]:
            for error in validation["errors"]:
                flash(error, "warning")
            return _render_quote_form(
                project,
                quote_from_form(request.form, fallback_quote=quote),
                quotes,
                field_errors=validation["field_errors"],
                quote_id=quote_id,
            )
        quote.update({
            "quote_number": validation["quote_number"] or quote["quote_number"],
            "quote_type": validation["quote_type"],
            "version": validation["version"],
            "date": validation["date"],
            "valid_until": validation["valid_until"],
            "items": validation["items"],
            "subtotal": validation["subtotal"],
            "tax_rate": validation["tax_rate"],
            "tax": round(validation["subtotal"] * validation["tax_rate"] / 100, 2),
            "total": round(validation["subtotal"] + validation["subtotal"] * validation["tax_rate"] / 100, 2),
            "currency": validation["currency"],
            "notes": validation["notes"],
            "project_basis_note": validation["project_basis_note"] if validation["quote_type"] == "Extraordinaria" else "",
            "specs": validation["specs"],
        })
        save("quotes", quotes)
        flash("Cotización actualizada.", "success")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))
    hydrated = hydrate_quote(quote, *catalog_maps())
    return _render_quote_form(project, hydrated, quotes, quote_id=quote_id)


@bp.route("/projects/<project_id>/quote/<quote_id>/view", endpoint="view_quote")
def view_quote(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        if project:
            return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())
    return render_template("quote_project_detail.html", project=project, quote=hydrated)


@bp.route("/projects/<project_id>/quote/<quote_id>/delete", methods=["POST"], endpoint="delete_quote")
def delete_quote(project_id, quote_id):
    save("quotes", [item for item in load("quotes") if item["id"] != quote_id])
    flash("Cotización eliminada.", "warning")
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")


@bp.route("/projects/<project_id>/quote/<quote_id>/approve", methods=["POST"], endpoint="approve_quote")
@admin_required
def approve_quote_route(project_id, quote_id):
    """Aprueba/activa una cotización.

    - General/Preliminar: marca la seleccionada como 'active' y las demás del
      mismo proyecto (mismos tipos) como 'obsolete'.
    - Extraordinaria: toggle active ↔ obsolete independiente.
    """
    quotes = load("quotes")
    changed = approve_quote(quote_id, quotes)
    if changed:
        save("quotes", quotes)
        flash("Estado de cotización actualizado.", "success")
    else:
        flash("Cotización no encontrada.", "danger")
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")


@bp.route(
    "/projects/<project_id>/quote/<quote_id>/purge-deleted-catalog",
    methods=["POST"],
    endpoint="purge_quote_deleted_catalog_items",
)
def purge_quote_deleted_catalog_items(project_id, quote_id):
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id and item["project_id"] == project_id), None)
    if not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    updated_quote, removed = purge_deleted_catalog_items_from_record(quote)
    subtotal = round(
        sum(safe_float(item.get("qty", 0)) * safe_float(item.get("price", 0)) for item in updated_quote.get("items", [])),
        2,
    )
    tax_rate = safe_float(updated_quote.get("tax_rate", 16), 16)
    updated_quote["subtotal"] = subtotal
    updated_quote["tax"] = round(subtotal * tax_rate / 100, 2)
    updated_quote["total"] = round(subtotal + updated_quote["tax"], 2)
    quote.update(updated_quote)
    save("quotes", quotes)
    flash(f"Se eliminaron {removed} partida(s) con catálogo eliminado de la cotización.", "warning")
    return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")


@bp.route("/projects/<project_id>/quote/<quote_id>/pdf", endpoint="quote_pdf")
def quote_pdf(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())
    pdf_name = f"{hydrated.get('quote_number', 'COT')}.pdf"
    inline = request.args.get("inline") == "1"
    try:
        pdf_bytes = build_quote_pdf(project, hydrated)
        return send_file(
            BytesIO(pdf_bytes),
            as_attachment=not inline,
            download_name=pdf_name,
            mimetype="application/pdf",
        )
    except Exception as exc:
        flash(f"Error al generar PDF: {exc}", "danger")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")


@bp.route("/projects/<project_id>/quote/<quote_id>/excel", endpoint="quote_excel")
def quote_excel(project_id, quote_id):
    """Genera la cotización como Excel y la sirve como descarga directa."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        flash(f"openpyxl no está instalado: {exc}", "danger")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")

    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))

    try:
        wb, filename = _build_quote_workbook(project, quote, Workbook, Alignment, Font)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        flash(f"Error al generar Excel: {type(exc).__name__}: {exc}", "danger")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")


def _build_quote_workbook(project, quote, Workbook, Alignment, Font):
    """Construye el workbook Excel de la cotización.

    Devuelve (wb, filename) para que el caller decida cómo persistirlo.
    """
    hydrated = hydrate_quote(quote, *catalog_maps())
    sections = hydrated.get("sections", [])
    has_sections = any(section.get("name") for section in sections)
    filename = f"{hydrated.get('quote_number', 'cotizacion')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 5    # #
    if has_sections:
        ws.column_dimensions["B"].width = 24   # Sección
        ws.column_dimensions["C"].width = 40   # Nombre / Descripción
        ws.column_dimensions["D"].width = 10   # Unidad
        ws.column_dimensions["E"].width = 12   # Cantidad
        ws.column_dimensions["F"].width = 16   # Precio unitario
        ws.column_dimensions["G"].width = 16   # Total
    else:
        ws.column_dimensions["B"].width = 40   # Nombre / Descripción
        ws.column_dimensions["C"].width = 10   # Unidad
        ws.column_dimensions["D"].width = 12   # Cantidad
        ws.column_dimensions["E"].width = 16   # Precio unitario
        ws.column_dimensions["F"].width = 16   # Total

    # ── Encabezado informativo ──────────────────────────────────────────────
    ws.append(["Cotización:", hydrated.get("quote_number", "")])
    ws.append(["Cliente:", hydrated.get("client", "")])
    ws.append(["Proyecto:", hydrated.get("project_name", "")])
    ws.append(["Fecha:", hydrated.get("date", "")])
    ws.append(["Moneda:", hydrated.get("currency", "")])
    ws.append([])  # fila vacía

    # ── Cabecera de tabla ───────────────────────────────────────────────────
    header_row = (
        ["#", "Sección", "Nombre", "Unidad", "Cantidad", "Precio unitario", "Total"]
        if has_sections
        else ["#", "Nombre", "Unidad", "Cantidad", "Precio unitario", "Total"]
    )
    ws.append(header_row)
    hdr_idx = ws.max_row
    total_col = 7 if has_sections else 6
    price_col = 6 if has_sections else 5
    qty_col = 5 if has_sections else 4
    section_label_col = 3 if has_sections else 2
    subtotal_label_col = 6 if has_sections else 5
    for col in range(1, total_col + 1):
        cell = ws.cell(row=hdr_idx, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ── Artículos ───────────────────────────────────────────────────────────
    row_num = 0
    for section in sections:
        if section.get("name"):
            section_row = [""] * total_col
            section_row[section_label_col - 1] = section["name"]
            ws.append(section_row)
            sect_cell = ws.cell(row=ws.max_row, column=section_label_col)
            sect_cell.font = Font(bold=True)

        for item in section.get("items", []):
            row_num += 1
            qty = item.get("qty", 0)
            price = item.get("price", 0)
            total = item.get("total", 0)
            if has_sections:
                ws.append([
                    row_num,
                    section.get("name", ""),
                    item.get("description", ""),
                    item.get("unit", ""),
                    qty,
                    price,
                    total,
                ])
            else:
                ws.append([
                    row_num,
                    item.get("description", ""),
                    item.get("unit", ""),
                    qty,
                    price,
                    total,
                ])
            r = ws.max_row
            ws.cell(row=r, column=qty_col).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=price_col).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=total_col).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=price_col).number_format = "#,##0.00"
            ws.cell(row=r, column=total_col).number_format = "#,##0.00"

        if section.get("name"):
            subtotal_row = [""] * total_col
            subtotal_row[section_label_col - 1] = f"Subtotal {section['name']}"
            subtotal_row[total_col - 1] = section.get("subtotal", 0)
            ws.append(subtotal_row)
            sr = ws.max_row
            ws.cell(row=sr, column=section_label_col).font = Font(bold=True)
            ws.cell(row=sr, column=total_col).number_format = "#,##0.00"
            ws.cell(row=sr, column=total_col).alignment = Alignment(horizontal="right")

    # ── Fila vacía de separación ────────────────────────────────────────────
    ws.append([])

    # ── Totales ─────────────────────────────────────────────────────────────
    subtotal = hydrated.get("subtotal", 0)
    tax_rate = hydrated.get("tax_rate", 0)
    tax = hydrated.get("tax", 0)
    total = hydrated.get("total", 0)

    for label, value in [
        ("Subtotal", subtotal),
        (f"IVA ({tax_rate}%)", tax),
        ("TOTAL", total),
    ]:
        total_row = [""] * total_col
        total_row[subtotal_label_col - 1] = label
        total_row[total_col - 1] = value
        ws.append(total_row)
        r = ws.max_row
        ws.cell(row=r, column=subtotal_label_col).font = Font(bold=True)
        ws.cell(row=r, column=subtotal_label_col).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=total_col).number_format = "#,##0.00"
        ws.cell(row=r, column=total_col).alignment = Alignment(horizontal="right")
        if label == "TOTAL":
            ws.cell(row=r, column=subtotal_label_col).font = Font(bold=True)
            ws.cell(row=r, column=total_col).font = Font(bold=True)

    return wb, filename


@bp.route("/projects/<project_id>/quote/<quote_id>/item/<int:item_index>/restore", methods=["POST"], endpoint="restore_deleted_item")
def restore_deleted_item(project_id, quote_id, item_index):
    """Restore a deleted catalog item by reconnecting it to a new catalog item"""
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id), None)

    if not project or not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))

    new_catalog_id = request.form.get("new_catalog_id", "").strip()
    if not new_catalog_id:
        flash("Debe seleccionar un nuevo artículo del catálogo.", "warning")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))

    from ..deletions import restore_deleted_catalog_item_in_record
    updated_quote, success = restore_deleted_catalog_item_in_record(quote, item_index, new_catalog_id)

    if success:
        # Update the quote in the list
        for i, q in enumerate(quotes):
            if q["id"] == quote_id:
                quotes[i] = updated_quote
                break
        save("quotes", quotes)
        flash("Partida reconectada exitosamente al nuevo artículo del catálogo.", "success")
    else:
        flash("No se pudo reconectar la partida. Verifique que existe y está eliminada.", "warning")

    return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))


@bp.route("/projects/<project_id>/quote/<quote_id>/item/<int:item_index>/preserve", methods=["POST"], endpoint="preserve_deleted_item")
def preserve_deleted_item(project_id, quote_id, item_index):
    """Mark a deleted catalog item as preserved (keep historical reference)"""
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id), None)

    if not project or not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))

    from ..deletions import preserve_deleted_catalog_item_in_record
    updated_quote, success = preserve_deleted_catalog_item_in_record(quote, item_index)

    if success:
        # Update the quote in the list
        for i, q in enumerate(quotes):
            if q["id"] == quote_id:
                quotes[i] = updated_quote
                break
        save("quotes", quotes)
        flash("Partida marcada como conservada históricamente.", "info")
    else:
        flash("No se pudo marcar la partida como conservada.", "warning")

    return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))


@bp.route("/projects/<project_id>/quote/<quote_id>/item/<int:item_index>/purge", methods=["POST"], endpoint="purge_deleted_item")
def purge_deleted_item(project_id, quote_id, item_index):
    """Purge a deleted catalog item from the quote"""
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id), None)

    if not project or not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))

    from ..deletions import purge_deleted_catalog_items_from_record

    # Create a temporary record with just this item to purge
    temp_record = {"items": [quote["items"][item_index]]}
    purged_record, purged_count = purge_deleted_catalog_items_from_record(temp_record)

    if purged_count > 0:
        # Remove the item from the quote
        updated_items = list(quote["items"])
        updated_items.pop(item_index)
        quote["items"] = updated_items

        # Recalculate totals
        from ..catalog import hydrate_quote
        hydrated = hydrate_quote(quote, *catalog_maps())
        quote.update({
            "subtotal": hydrated["subtotal"],
            "tax": hydrated["tax"],
            "total": hydrated["total"]
        })

        # Update the quote in the list
        for i, q in enumerate(quotes):
            if q["id"] == quote_id:
                quotes[i] = quote
                break
        save("quotes", quotes)
        flash("Partida eliminada definitivamente de la cotización.", "warning")
    else:
        flash("No se pudo eliminar la partida.", "warning")

    return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=quote_id))


@bp.route("/projects/<project_id>/quote/<quote_id>/resumen", endpoint="quote_resumen")
def quote_resumen(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())
    agg_items = aggregate_quote_items(hydrated["items"])
    subtotal = round(sum(safe_float(i.get("total", 0)) for i in agg_items), 2)
    tax_rate = safe_float(hydrated.get("tax_rate", 16), 16)
    resumen = dict(hydrated)
    resumen["items"] = agg_items
    resumen["subtotal"] = subtotal
    resumen["tax"] = round(subtotal * tax_rate / 100, 2)
    resumen["total"] = round(subtotal + resumen["tax"], 2)
    return render_template("quote_resumen.html", project=project, quote=resumen)


@bp.route("/projects/<project_id>/quote/<quote_id>/resumen/pdf", endpoint="quote_resumen_pdf")
def quote_resumen_pdf(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())
    agg_items = aggregate_quote_items(hydrated["items"])
    subtotal = round(sum(safe_float(i.get("total", 0)) for i in agg_items), 2)
    tax_rate = safe_float(hydrated.get("tax_rate", 16), 16)
    resumen = dict(hydrated)
    resumen["items"] = agg_items
    resumen["subtotal"] = subtotal
    resumen["tax"] = round(subtotal * tax_rate / 100, 2)
    resumen["total"] = round(subtotal + resumen["tax"], 2)
    pdf_name = f"{hydrated.get('quote_number', 'COT')}-Resumen.pdf"
    try:
        pdf_bytes = build_quote_pdf(project, resumen)
        return send_file(BytesIO(pdf_bytes), as_attachment=True, download_name=pdf_name, mimetype="application/pdf")
    except Exception as exc:
        flash(f"Error al generar PDF: {exc}", "danger")
        return redirect(url_for("quotes_bp.quote_resumen", project_id=project_id, quote_id=quote_id))


@bp.route("/projects/<project_id>/quote/<quote_id>/resumen/excel", endpoint="quote_resumen_excel")
def quote_resumen_excel(project_id, quote_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        flash(f"openpyxl no está instalado: {exc}", "danger")
        return redirect(url_for("quotes_bp.quote_resumen", project_id=project_id, quote_id=quote_id))
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())
    agg_items = aggregate_quote_items(hydrated["items"])
    subtotal = round(sum(safe_float(i.get("total", 0)) for i in agg_items), 2)
    tax_rate = safe_float(hydrated.get("tax_rate", 16), 16)
    resumen = dict(hydrated)
    resumen["items"] = agg_items
    resumen["subtotal"] = subtotal
    resumen["tax"] = round(subtotal * tax_rate / 100, 2)
    resumen["total"] = round(subtotal + resumen["tax"], 2)
    try:
        wb, filename = _build_quote_workbook(project, resumen, Workbook, Alignment, Font)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        base = filename.rsplit(".", 1)[0]
        return send_file(buf, as_attachment=True, download_name=f"{base}-Resumen.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as exc:
        flash(f"Error al generar Excel: {type(exc).__name__}: {exc}", "danger")
        return redirect(url_for("quotes_bp.quote_resumen", project_id=project_id, quote_id=quote_id))


@bp.route("/projects/<project_id>/quote/<quote_id>/csv", endpoint="quote_csv_export")
def quote_csv_export(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quote = next((item for item in load("quotes") if item["id"] == quote_id), None)
    if not project or not quote or quote.get("project_id") != project_id:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))
    hydrated = hydrate_quote(quote, *catalog_maps())

    def _csv_escape(value):
        return str(value or "").replace('"', '""')

    lines = []
    lines.append(f'#proyecto_clave,"{_csv_escape(project.get("clave", ""))}"')
    lines.append(f'#quote_type,"{_csv_escape(hydrated.get("quote_type", "General"))}"')
    lines.append(f'#fecha,"{_csv_escape(hydrated.get("date", ""))}"')
    lines.append(f'#version,"{_csv_escape(hydrated.get("version", ""))}"')
    lines.append(f'#currency,"{_csv_escape(hydrated.get("currency", "MXN"))}"')
    lines.append("description,unit,qty,price,section")
    for item in hydrated.get("items", []):
        if item.get("kind") == "section":
            lines.append(f'"","","","","{_csv_escape(item.get("section"))}"')
            continue
        lines.append(
            f'"{_csv_escape(item.get("description"))}",'
            f'"{_csv_escape(item.get("unit"))}",'
            f'{item.get("qty", 0)},'
            f'{item.get("price", 0)},'
            f'"{_csv_escape(item.get("section"))}"'
        )
    csv_content = "﻿" + "\n".join(lines) + "\n"
    safe_name = "".join(c for c in hydrated.get("quote_number", "cotizacion") if c not in r'"\/:<>|?*')
    filename = f"{safe_name}.csv"
    return Response(
        csv_content,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@bp.route("/projects/<project_id>/quote/<quote_id>/duplicate", methods=["POST"], endpoint="quote_duplicate")
def quote_duplicate(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    source = next((item for item in quotes if item["id"] == quote_id and item.get("project_id") == project_id), None)
    if not project or not source:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("project_detail", project_id=project_id) + "#tab-quote")
    new_quote = copy.deepcopy(source)
    new_quote["id"] = new_id()
    base_number = source.get("quote_number", "COT")
    existing_numbers = {q.get("quote_number", "") for q in quotes}
    candidate = f"{base_number}-Copia"
    counter = 2
    while candidate in existing_numbers:
        candidate = f"{base_number}-Copia{counter}"
        counter += 1
    new_quote["quote_number"] = candidate
    new_quote["approval_status"] = "draft"
    new_quote["created_at"] = today()
    new_quote["valid_until"] = ""
    new_quote.pop("csv_origen", None)
    new_quote.pop("csv_sources", None)
    new_quote.pop("csv_metadata", None)
    quotes.append(new_quote)
    save("quotes", quotes)
    flash(f"Cotización duplicada como {candidate}.", "success")
    return redirect(url_for("quotes_bp.edit_quote", project_id=project_id, quote_id=new_quote["id"]))


@bp.route("/audit/deleted-catalog", endpoint="audit_deleted_catalog")
def audit_deleted_catalog():
    """Audit all quotes and LDMs for deleted catalog items"""
    from ..deletions import audit_deleted_catalog_items

    quotes = load("quotes")
    ldms = load("materiales")

    quote_audit = audit_deleted_catalog_items(quotes, "quote")
    ldm_audit = audit_deleted_catalog_items(ldms, "ldm")

    # Combine results
    combined_audit = {
        "total_quotes": quote_audit["total_records"],
        "total_ldms": ldm_audit["total_records"],
        "quotes_with_deleted_items": quote_audit["records_with_deleted_items"],
        "ldms_with_deleted_items": ldm_audit["records_with_deleted_items"],
        "total_deleted_items": quote_audit["total_deleted_items"] + ldm_audit["total_deleted_items"],
        "preserved_items": quote_audit["preserved_items"] + ldm_audit["preserved_items"],
        "unresolved_items": quote_audit["unresolved_items"] + ldm_audit["unresolved_items"],
        "details": quote_audit["details"] + ldm_audit["details"]
    }

    return render_template("audit_deleted_catalog.html", audit=combined_audit)



@bp.route("/projects/<project_id>/quote/<quote_id>/pdf-editor", methods=["GET", "POST"], endpoint="quote_pdf_editor")
def quote_pdf_editor(project_id, quote_id):
    project = next((item for item in load("projects") if item["id"] == project_id), None)
    quotes = load("quotes")
    quote = next((item for item in quotes if item["id"] == quote_id), None)
    if not project or not quote:
        flash("Cotización no encontrada.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        specs = dict(quote.get("specs") or {})
        for field in ("condiciones_pago", "exclusiones", "validez", "forma_entrega", "contacto"):
            specs[field] = request.form.get(field, "").strip()
        specs["alcance_custom"] = request.form.get("alcance_custom", "").strip()
        specs["nota_precio"] = request.form.get("nota_precio", "").strip()
        specs["portada_spacing"] = request.form.get("portada_spacing", "").strip()
        quote["specs"] = specs
        quote["notes"] = request.form.get("notes", "").strip()
        basis = request.form.get("project_basis_note", "").strip()
        if quote_type_key(quote.get("quote_type")) == "Extraordinaria":
            quote["project_basis_note"] = basis
        else:
            specs["basis_note_override"] = basis
        save("quotes", quotes)
        flash("Cambios guardados en el editor PDF.", "success")
        return redirect(url_for("quotes_bp.quote_pdf_editor", project_id=project_id, quote_id=quote_id))

    from ..pdfs import quote_cover_copy, quote_scope_paragraphs, quote_terms
    from ..company_config import get_company

    hydrated = hydrate_quote(quote, *catalog_maps())
    cover_title, cover_subtitle = quote_cover_copy(hydrated)
    qt = quote_type_key(quote.get("quote_type"))
    specs = quote.get("specs") or {}
    if qt == "Extraordinaria":
        basis_note_edit = quote.get("project_basis_note") or ""
    else:
        basis_note_edit = specs.get("basis_note_override") or ""

    company = get_company()
    logo_rel = company.get("logo") or ""
    import os
    from ..storage import BASE_DIR
    logo_url = None
    if logo_rel:
        logo_abs = os.path.join(BASE_DIR, "static", logo_rel)
        if os.path.isfile(logo_abs):
            logo_url = "/static/" + logo_rel
    company_info_parts = [p for p in [company.get("address", ""), company.get("rut", "")] if p]
    company_info = "  ·  ".join(company_info_parts)

    return render_template(
        "quote_pdf_editor.html",
        project=project,
        quote=hydrated,
        cover_title=cover_title,
        cover_subtitle=cover_subtitle,
        basis_note_edit=basis_note_edit,
        specs=specs,
        default_scope=quote_scope_paragraphs(),
        default_terms=quote_terms(),
        quote_type=qt,
        company_name=company.get("name") or "Project Tracker",
        company_info=company_info,
        logo_url=logo_url,
    )
